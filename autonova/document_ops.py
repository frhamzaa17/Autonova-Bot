from __future__ import annotations

import csv
import html
import re
import zipfile
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape

from .config import GENERATED_DIR, ensure_directories


def safe_name(text: str, suffix: str) -> Path:
    ensure_directories()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()[:40] or "document"
    return GENERATED_DIR / f"{stamp}_{stem}.{suffix}"


def draft_text_document(prompt: str, knowledge: list[dict]) -> Path:
    title = "AutoNova Draft"
    context = "\n".join(f"- {item['title']}: {item['content']}" for item in knowledge[:3])
    body = f"""{title}

Request:
{prompt}

Relevant business context:
{context or "No matching knowledge-base entries found."}

Draft:
This document is prepared for the real-estate business based on the request above.
Please review names, amounts, dates, property identifiers, payment terms, and legal clauses before client use.

Key points:
- Parties and property details to be confirmed by operations.
- Pricing, brokerage, taxes, and payment milestones to be verified from the latest sheet.
- Final version should be reviewed by the authorized business owner.
"""
    path = safe_name("autonova_draft", "txt")
    path.write_text(body, encoding="utf-8")
    return path


def draft_docx(prompt: str, knowledge: list[dict]) -> Path:
    path = safe_name("autonova_draft", "docx")
    paragraphs = [
        "AutoNova Draft",
        "",
        f"Request: {prompt}",
        "",
        "Relevant business context:",
    ]
    paragraphs += [f"- {item['title']}: {item['content']}" for item in knowledge[:3]]
    paragraphs += [
        "",
        "Draft:",
        "This document is prepared for the real-estate business based on the request above.",
        "Please verify names, amounts, dates, property identifiers, and legal clauses before client use.",
    ]
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>'
        + "".join(f"<w:p><w:r><w:t>{escape(p)}</w:t></w:r></w:p>" for p in paragraphs)
        + "</w:body></w:document>"
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", '<?xml version="1.0" encoding="UTF-8"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/></Types>')
        zf.writestr("_rels/.rels", '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/></Relationships>')
        zf.writestr("word/document.xml", document_xml)
    return path


def draft_pdf(prompt: str, knowledge: list[dict]) -> Path:
    path = safe_name("autonova_draft", "pdf")
    context = "\n".join(f"- {item['title']}: {item['content']}" for item in knowledge[:3])
    text = f"""AutoNova Draft

Request:
{prompt}

Relevant business context:
{context or "No matching knowledge-base entries found."}

Draft:
This document is prepared for the real-estate business based on the request above.
Please verify names, amounts, dates, property identifiers, and legal clauses before client use.
"""
    lines = []
    for raw_line in text.splitlines():
        while len(raw_line) > 88:
            lines.append(raw_line[:88])
            raw_line = raw_line[88:]
        lines.append(raw_line)
    content = ["BT", "/F1 11 Tf", "50 780 Td", "14 TL"]
    for index, line in enumerate(lines[:48]):
        safe = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        if index:
            content.append("T*")
        content.append(f"({safe}) Tj")
    content.append("ET")
    stream = "\n".join(content).encode("latin-1", errors="replace")
    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n",
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n",
        b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n",
        f"5 0 obj << /Length {len(stream)} >> stream\n".encode("ascii") + stream + b"\nendstream endobj\n",
    ]
    out = [b"%PDF-1.4\n"]
    offsets = [0]
    for obj in objects:
        offsets.append(sum(len(part) for part in out))
        out.append(obj)
    xref_start = sum(len(part) for part in out)
    xref = [b"xref\n0 6\n", b"0000000000 65535 f \n"]
    for offset in offsets[1:]:
        xref.append(f"{offset:010d} 00000 n \n".encode("ascii"))
    out.extend(xref)
    out.append(b"trailer << /Size 6 /Root 1 0 R >>\n")
    out.append(f"startxref\n{xref_start}\n%%EOF".encode("ascii"))
    path.write_bytes(b"".join(out))
    return path


def summarize_text_file(path: Path) -> str:
    text = path.read_text(encoding="utf-8", errors="ignore")
    compact = re.sub(r"\s+", " ", text).strip()
    return compact[:1000] + ("..." if len(compact) > 1000 else "")


def extract_document_text(path: Path, max_chars: int = 12000) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        text = path.read_text(encoding="utf-8", errors="ignore")
    elif suffix == ".pdf":
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise RuntimeError("Install dependencies with `pip install -r requirements.txt` to read PDF uploads.") from exc

        reader = PdfReader(path)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    elif suffix == ".docx":
        try:
            import docx2txt
        except ImportError as exc:
            raise RuntimeError("Install dependencies with `pip install -r requirements.txt` to read DOCX uploads.") from exc

        text = docx2txt.process(str(path)) or ""
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Install dependencies with `pip install -r requirements.txt` to read XLSX uploads.") from exc

        workbook = load_workbook(path, data_only=False)
        rows = []
        for sheet in workbook.worksheets:
            rows.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value is not None]
                if values:
                    rows.append(" | ".join(values))
        text = "\n".join(rows)
    else:
        return ""
    compact = re.sub(r"\s+", " ", text).strip()
    return compact[:max_chars]


def edit_csv(path: Path, instruction: str) -> Path:
    rows = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = list(reader.fieldnames or [])
        for row in reader:
            rows.append(row)

    if not fieldnames:
        raise ValueError("CSV has no header row.")

    lower = instruction.lower()
    if "mark" in lower and "status" not in [f.lower() for f in fieldnames]:
        fieldnames.append("status")
    if "review" in lower and "status" in [f.lower() for f in fieldnames]:
        status_col = next(f for f in fieldnames if f.lower() == "status")
        for row in rows:
            row[status_col] = "Review"
    if "total" in lower:
        numeric_cols = []
        for name in fieldnames:
            try:
                sum(float(row.get(name, "") or 0) for row in rows)
                numeric_cols.append(name)
            except ValueError:
                pass
        if numeric_cols:
            total_row = {name: "" for name in fieldnames}
            total_row[fieldnames[0]] = "TOTAL"
            for name in numeric_cols:
                total_row[name] = str(sum(float(row.get(name, "") or 0) for row in rows))
            rows.append(total_row)

    output = safe_name(path.stem + "_edited", "csv")
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return output


def edit_xlsx(path: Path, instruction: str) -> Path:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Install dependencies with `pip install -r requirements.txt` to edit .xlsx spreadsheets.") from exc

    wb = load_workbook(path)
    ws = wb.active
    lower = instruction.lower()
    headers = [str(cell.value or "").strip() for cell in ws[1]]

    if "mark" in lower and "status" not in [h.lower() for h in headers]:
        status_col = len(headers) + 1
        ws.cell(row=1, column=status_col).value = "status"
    else:
        status_col = next((i + 1 for i, h in enumerate(headers) if h.lower() == "status"), None)

    if "review" in lower and status_col:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=status_col).value = "Review"

    if "total" in lower or "formula" in lower:
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1).value = "TOTAL"
        for col in range(2, ws.max_column + 1):
            values = [ws.cell(row=row, column=col).value for row in range(2, ws.max_row + 1)]
            if any(isinstance(value, (int, float)) for value in values):
                letter = ws.cell(row=1, column=col).column_letter
                ws.cell(row=total_row, column=col).value = f"=SUM({letter}2:{letter}{total_row - 1})"

    output = safe_name(path.stem + "_edited", "xlsx")
    wb.save(output)
    return output


def html_escape(text: str) -> str:
    return html.escape(text, quote=True)
