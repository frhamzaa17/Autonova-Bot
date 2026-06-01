from __future__ import annotations

import re
from pathlib import Path
from dataclasses import dataclass

from langchain_chroma import Chroma
from langchain_community.document_loaders import Docx2txtLoader, PyPDFLoader, TextLoader
from langchain_core.documents import Document
from langchain_ollama import OllamaEmbeddings
from langchain_text_splitters import RecursiveCharacterTextSplitter

from utils.config import Settings, load_settings, safe_workspace_id
from rag.ocr import IMAGE_EXTENSIONS, is_image_file, extract_image_text


SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx"} | IMAGE_EXTENSIONS
TOKEN_RE = re.compile(r"[a-z0-9]+")


@dataclass(frozen=True)
class KnowledgeFile:
    path: Path
    file_name: str
    document_type: str
    record_count: int


@dataclass(frozen=True)
class KnowledgeRemovalResult:
    removed_files: list[Path]
    structured_removed: int
    vectors_removed: bool


def _tenant_roots(settings: Settings, tenant_id: str | None, include_generated: bool = False) -> list[Path]:
    if tenant_id:
        workspace = safe_workspace_id(tenant_id)
        roots = [settings.uploads_dir / workspace, settings.data_dir / workspace]
        if include_generated:
            roots.append(settings.generated_dir / workspace)
        return roots
    roots = [settings.uploads_dir, settings.data_dir]
    if include_generated:
        roots.append(settings.generated_dir)
    return roots


def _embeddings(settings: Settings) -> OllamaEmbeddings:
    return OllamaEmbeddings(model=settings.ollama_embedding_model, base_url=settings.ollama_url)


def _load_file(path: Path) -> list[Document]:
    suffix = path.suffix.lower()
    if is_image_file(path):
        text = extract_image_text(path)
        return [Document(page_content=text, metadata={"source": str(path), "kind": "ocr_image"})] if text else []
    if suffix in {".txt", ".md"}:
        return TextLoader(str(path), encoding="utf-8").load()
    if suffix == ".pdf":
        return PyPDFLoader(str(path)).load()
    if suffix == ".docx":
        return Docx2txtLoader(str(path)).load()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=False)
        docs = []
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(value) for value in row if value is not None]
                if values:
                    rows.append(" | ".join(values))
            if rows:
                docs.append(Document(page_content="\n".join(rows), metadata={"source": str(path), "sheet": sheet.title}))
        return docs
    return []


def _collection_name(tenant_id: str | None = None) -> str:
    if not tenant_id:
        return "local_knowledge"
    safe = "".join(char if char.isalnum() or char in {"_", "-"} else "_" for char in tenant_id)
    return f"local_knowledge_{safe}"[:63]


def _vector_store(settings: Settings, tenant_id: str | None = None) -> Chroma:
    return Chroma(
        collection_name=_collection_name(tenant_id),
        persist_directory=str(settings.chroma_dir),
        embedding_function=_embeddings(settings),
    )


def _tokens(text: str) -> set[str]:
    tokens = set(TOKEN_RE.findall(text.lower()))
    if {"property", "properties", "prop"} & tokens:
        tokens.update({"property", "properties", "prop", "listing", "listings", "portfolio"})
    if {"tenant", "tenants", "rent", "rental"} & tokens:
        tokens.update({"tenant", "tenants", "rent", "rental", "agreement", "lease"})
    if {"client", "clients", "contact", "contacts"} & tokens:
        tokens.update({"client", "clients", "contact", "contacts", "directory"})
    return tokens


def _candidate_files(settings: Settings, tenant_id: str | None = None) -> list[Path]:
    files: list[Path] = []
    for root in _tenant_roots(settings, tenant_id):
        if not root.exists():
            continue
        paths = root.rglob("*") if tenant_id else root.glob("*")
        for path in paths:
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
                files.append(path)
    return files


def _delete_vectors_for_source(path: Path, tenant_id: str | None) -> bool:
    settings = load_settings()
    store = _vector_store(settings, tenant_id)
    deleted = False
    sources = {str(path), str(path.resolve())}
    for source in sources:
        try:
            store._collection.delete(where={"source": source})
            deleted = True
        except Exception:
            pass
    return deleted


def _clear_vectors(tenant_id: str | None) -> bool:
    settings = load_settings()
    name = _collection_name(tenant_id)
    store = _vector_store(settings, tenant_id)
    try:
        store._client.delete_collection(name)
        return True
    except Exception:
        pass
    try:
        existing = store._collection.get(include=[])
        ids = existing.get("ids", [])
        if ids:
            store._collection.delete(ids=ids)
        return True
    except Exception:
        return False


def knowledge_files(tenant_id: str | None) -> list[KnowledgeFile]:
    from rag.structured_store import structured_documents

    files: list[KnowledgeFile] = []
    for document in structured_documents(tenant_id):
        source = document.get("source_file")
        if not source:
            continue
        path = Path(source)
        files.append(
            KnowledgeFile(
                path=path,
                file_name=document.get("file_name") or path.name,
                document_type=document.get("document_type", "unknown"),
                record_count=int(document.get("summary", {}).get("record_count") or 0),
            )
        )
    return files


def knowledge_summary(tenant_id: str | None) -> str:
    files = knowledge_files(tenant_id)
    if not files:
        return "Your knowledge base is empty."
    types: dict[str, int] = {}
    for item in files:
        types[item.document_type] = types.get(item.document_type, 0) + 1
    type_summary = ", ".join(f"{kind}: {count}" for kind, count in sorted(types.items()))
    names = "\n".join(f"- {item.file_name} ({item.document_type}, {item.record_count} records)" for item in files[:30])
    more = f"\n...and {len(files) - 30} more." if len(files) > 30 else ""
    return f"Your knowledge base has {len(files)} file(s). Types: {type_summary}.\n{names}{more}"


def _matches_file(item: KnowledgeFile, selector: str) -> bool:
    needle = selector.strip().lower().strip("\"'")
    if not needle:
        return False
    name = item.file_name.lower()
    stem = Path(item.file_name).stem.lower()
    return needle == name or needle == stem or needle in name or needle in stem


def remove_knowledge_file(selector: str, tenant_id: str | None) -> KnowledgeRemovalResult:
    from rag.structured_store import remove_structured_document

    matches = [item for item in knowledge_files(tenant_id) if _matches_file(item, selector)]
    removed_files: list[Path] = []
    structured_removed = 0
    vectors_removed = False
    settings = load_settings()
    roots = [root.resolve() for root in _tenant_roots(settings, tenant_id, include_generated=True) if root.exists()]
    for item in matches:
        path = item.path
        try:
            resolved = path.resolve()
        except OSError:
            resolved = path
        if not any(resolved == root or root in resolved.parents for root in roots):
            continue
        if remove_structured_document(path, tenant_id):
            structured_removed += 1
        vectors_removed = _delete_vectors_for_source(path, tenant_id) or vectors_removed
        if path.exists() and path.is_file():
            path.unlink()
            removed_files.append(path)
    return KnowledgeRemovalResult(removed_files, structured_removed, vectors_removed)


def clear_knowledge_base(tenant_id: str | None) -> KnowledgeRemovalResult:
    from rag.structured_store import clear_structured_documents

    settings = load_settings()
    removed_files: list[Path] = []
    for root in _tenant_roots(settings, tenant_id, include_generated=True):
        if not root.exists():
            continue
        paths = root.rglob("*") if tenant_id else root.glob("*")
        for path in paths:
            if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
                path.unlink()
                removed_files.append(path)
    structured_removed = clear_structured_documents(tenant_id)
    vectors_removed = _clear_vectors(tenant_id)
    return KnowledgeRemovalResult(removed_files, structured_removed, vectors_removed)


def _lexical_context(query: str, settings: Settings, tenant_id: str | None = None, limit: int = 3, max_chars: int = 2200) -> str:
    query_tokens = _tokens(query)
    if not query_tokens:
        return ""

    scored: list[tuple[int, Path, str]] = []
    for path in _candidate_files(settings, tenant_id):
        try:
            docs = _load_file(path)
        except Exception:
            continue
        text = "\n".join(doc.page_content for doc in docs if doc.page_content.strip())
        if not text:
            continue
        haystack = f"{path.stem} {path.name} {text}"
        haystack_tokens = _tokens(haystack)
        score = len(query_tokens & haystack_tokens)
        filename_score = len(query_tokens & _tokens(path.stem)) * 3
        if score or filename_score:
            scored.append((score + filename_score, path, text[:max_chars]))

    scored.sort(key=lambda item: item[0], reverse=True)
    return "\n\n".join(
        f"Knowledge excerpt:\n{text}"
        for _score, path, text in scored[:limit]
    )


def ingest_data_folder(settings: Settings | None = None, tenant_id: str | None = None) -> int:
    settings = settings or load_settings()
    docs: list[Document] = []
    root = settings.data_dir / safe_workspace_id(tenant_id) if tenant_id else settings.data_dir
    if not root.exists():
        return 0
    paths = root.rglob("*") if tenant_id else root.glob("*")
    for path in paths:
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
            try:
                from rag.structured_store import upsert_document

                upsert_document(path, tenant_id)
            except Exception:
                pass
            docs.extend(_load_file(path))

    if not docs:
        return 0

    splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=160)
    chunks = splitter.split_documents(docs)
    store = _vector_store(settings, tenant_id)
    store.add_documents(chunks)
    return len(chunks)


def ingest_file(path: Path, tenant_id: str) -> int:
    settings = load_settings()
    try:
        from rag.structured_store import upsert_document

        upsert_document(path, tenant_id)
    except Exception:
        pass
    docs = _load_file(path)
    if not docs:
        return 0
    splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=160)
    chunks = splitter.split_documents(docs)
    store = _vector_store(settings, tenant_id)
    store.add_documents(chunks)
    return len(chunks)


def retrieve_context(query: str, min_relevance: float = 0.25, k: int = 6, tenant_id: str | None = None) -> tuple[str, bool]:
    settings = load_settings()
    store = _vector_store(settings, tenant_id)
    contexts: list[str] = []
    try:
        results = store.similarity_search_with_relevance_scores(query, k=k)
    except Exception:
        results = []
    relevant = [(doc, score) for doc, score in results if score >= min_relevance]

    lexical = _lexical_context(query, settings, tenant_id)
    if lexical:
        contexts.append(lexical)

    if relevant:
        contexts.append(
            "\n\n".join(
                f"Knowledge excerpt:\n{doc.page_content}"
                for doc, _score in relevant
            )
        )

    context = "\n\n".join(contexts)
    return context, bool(context)
